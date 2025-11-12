#!/usr/bin/env python3
"""
RLID-NET Visualization and Reporting Module
Generates individual report files for training analysis and results
"""

import os
import logging
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
from typing import Dict, List, Any, Optional
from pathlib import Path
from datetime import datetime

try:
    import openpyxl
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.styles import Font, Alignment, PatternFill
except ImportError:
    print("Warning: openpyxl not available. Excel functionality will be limited.")

from .config import LID_TYPES, LID_COSTS
from ..rl.agent import TrainingMetrics


class RLIDVisualizer:
    """
    RLID-NET Visualization and Report Generator
    
    Creates individual visualization files for training analysis and results
    """
    
    def __init__(self, output_dir: str, logger: Optional[logging.Logger] = None):
        """
        Initialize visualizer
        
        Args:
            output_dir: Directory to save visualization files
            logger: Logger instance
        """
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.logger = logger or logging.getLogger(__name__)
        
        # Set up matplotlib and seaborn styles
        plt.style.use('default')
        sns.set_palette("husl")
        
        # Configure matplotlib for better plots
        plt.rcParams['figure.figsize'] = (10, 6)
        plt.rcParams['figure.dpi'] = 300
        plt.rcParams['savefig.dpi'] = 300
        plt.rcParams['savefig.bbox'] = 'tight'
        plt.rcParams['font.size'] = 10
        plt.rcParams['axes.grid'] = True
        plt.rcParams['grid.alpha'] = 0.3
        
        self.logger.info(f"RLID Visualizer initialized - Output directory: {output_dir}")
    
    def generate_all_reports(self, 
                           training_metrics: TrainingMetrics,
                           baseline_runoff: float,
                           final_lid_placements: List[Dict],
                           evaluation_results: Optional[Dict] = None) -> Dict[str, str]:
        """
        Generate all visualization reports
        
        Args:
            training_metrics: Training metrics from agent
            baseline_runoff: Baseline runoff value (m³)
            final_lid_placements: Final LID placement configuration
            evaluation_results: Optional evaluation results
            
        Returns:
            Dictionary with generated file paths
        """
        self.logger.info("Generating all visualization reports...")
        
        generated_files = {}
        
        try:
            # 1. Reward trend graph
            reward_file = self.create_reward_trend_graph(training_metrics)
            generated_files['reward_trend'] = str(reward_file)
            
            # 2. Loss trend graph
            loss_file = self.create_loss_trend_graph(training_metrics)
            generated_files['loss_trend'] = str(loss_file)
            
            # 3. Training metrics Excel
            metrics_file = self.create_training_metrics_excel(training_metrics)
            generated_files['training_metrics'] = str(metrics_file)
            
            # 4. LID placement summary Excel
            placement_file = self.create_lid_placement_summary_excel(final_lid_placements, baseline_runoff, training_metrics)
            generated_files['lid_placement_summary'] = str(placement_file)
            
            # 5. Baseline comparison chart
            comparison_file = self.create_baseline_comparison_chart(
                training_metrics, baseline_runoff, final_lid_placements
            )
            generated_files['baseline_comparison'] = str(comparison_file)
            
            self.logger.info("All visualization reports generated successfully!")
            
        except Exception as e:
            self.logger.error(f"Error generating reports: {str(e)}")
            raise
        
        return generated_files
    
    def create_reward_trend_graph(self, training_metrics: TrainingMetrics) -> Path:
        """
        Generate reward trend graph
        
        Args:
            training_metrics: Training metrics
            
        Returns:
            Path to generated file
        """
        self.logger.info("Creating reward trend graph...")
        
        fig, ax = plt.subplots(figsize=(12, 8))
        
        episodes = range(1, len(training_metrics.episode_rewards) + 1)
        rewards = training_metrics.episode_rewards
        
        # Plot reward trend
        ax.plot(episodes, rewards, linewidth=2, color='#2E86AB', alpha=0.7, label='Episode Reward')
        
        # Add moving average for smoother trend
        if len(rewards) > 10:
            window_size = max(5, len(rewards) // 20)
            moving_avg = pd.Series(rewards).rolling(window=window_size).mean()
            ax.plot(episodes, moving_avg, linewidth=3, color='#A23B72', label=f'Moving Average ({window_size} episodes)')
        
        # Customize plot
        ax.set_xlabel('Episode', fontsize=12, fontweight='bold')
        ax.set_ylabel('Reward', fontsize=12, fontweight='bold')
        ax.set_title('RLID-NET Training: Reward Progress Over Episodes', fontsize=14, fontweight='bold', pad=20)
        ax.legend(fontsize=11)
        ax.grid(True, alpha=0.3)
        
        # Add statistics text
        stats_text = f'Final Reward: {rewards[-1]:.2f}\n'
        stats_text += f'Best Reward: {max(rewards):.2f}\n'
        stats_text += f'Mean Reward: {np.mean(rewards):.2f}\n'
        stats_text += f'Std Reward: {np.std(rewards):.2f}'
        
        ax.text(0.02, 0.98, stats_text, transform=ax.transAxes, fontsize=10,
                verticalalignment='top', bbox=dict(boxstyle='round', facecolor='white', alpha=0.8))
        
        # Save figure
        output_file = self.output_dir / 'reward_trend.png'
        plt.savefig(output_file, bbox_inches='tight', facecolor='white')
        plt.close()
        
        self.logger.info(f"Reward trend graph saved: {output_file}")
        return output_file
    
    def create_loss_trend_graph(self, training_metrics: TrainingMetrics) -> Path:
        """
        Generate loss trend graph
        
        Args:
            training_metrics: Training metrics
            
        Returns:
            Path to generated file
        """
        self.logger.info("Creating loss trend graph...")
        
        fig, ax = plt.subplots(figsize=(12, 8))
        
        episodes = range(1, len(training_metrics.episode_losses) + 1)
        losses = training_metrics.episode_losses
        
        # Plot loss trend
        ax.plot(episodes, losses, linewidth=2, color='#F18F01', alpha=0.7, label='Episode Loss')
        
        # Add moving average for smoother trend
        if len(losses) > 10:
            window_size = max(5, len(losses) // 20)
            moving_avg = pd.Series(losses).rolling(window=window_size).mean()
            ax.plot(episodes, moving_avg, linewidth=3, color='#C73E1D', label=f'Moving Average ({window_size} episodes)')
        
        # Customize plot
        ax.set_xlabel('Episode', fontsize=12, fontweight='bold')
        ax.set_ylabel('Loss', fontsize=12, fontweight='bold')
        ax.set_title('RLID-NET Training: Loss Progress Over Episodes', fontsize=14, fontweight='bold', pad=20)
        ax.legend(fontsize=11)
        ax.grid(True, alpha=0.3)
        ax.set_yscale('log')  # Log scale for loss visualization
        
        # Add statistics text
        stats_text = f'Final Loss: {losses[-1]:.4f}\n'
        stats_text += f'Min Loss: {min(losses):.4f}\n'
        stats_text += f'Mean Loss: {np.mean(losses):.4f}\n'
        stats_text += f'Std Loss: {np.std(losses):.4f}'
        
        ax.text(0.02, 0.98, stats_text, transform=ax.transAxes, fontsize=10,
                verticalalignment='top', bbox=dict(boxstyle='round', facecolor='white', alpha=0.8))
        
        # Save figure
        output_file = self.output_dir / 'loss_trend.png'
        plt.savefig(output_file, bbox_inches='tight', facecolor='white')
        plt.close()
        
        self.logger.info(f"Loss trend graph saved: {output_file}")
        return output_file
    
    def create_training_metrics_excel(self, training_metrics: TrainingMetrics) -> Path:
        """
        Generate training metrics Excel file
        
        Args:
            training_metrics: Training metrics
            
        Returns:
            Path to generated file
        """
        self.logger.info("Creating training metrics Excel file...")
        
        # Prepare data
        episodes = list(range(1, len(training_metrics.episode_rewards) + 1))
        
        # Create DataFrame
        df = pd.DataFrame({
            'Episode': episodes,
            'Reward': training_metrics.episode_rewards,
            'Loss': training_metrics.episode_losses,
            'Total_Cost_KRW': training_metrics.episode_costs,
            'Runoff_Reduction_m3': training_metrics.episode_runoff_reductions,
            'Epsilon': training_metrics.epsilon_values
        })
        
        # Add calculated columns
        df['Cumulative_Best_Reward'] = df['Reward'].cummax()
        df['Reward_Moving_Avg_10'] = df['Reward'].rolling(window=10, min_periods=1).mean()
        df['Loss_Moving_Avg_10'] = df['Loss'].rolling(window=10, min_periods=1).mean()
        
        # Save to Excel
        output_file = self.output_dir / 'training_metrics.xlsx'
        
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # Main data sheet
            df.to_excel(writer, sheet_name='Training_Metrics', index=False)
            
            # Summary statistics sheet
            summary_stats = pd.DataFrame({
                'Metric': ['Total Episodes', 'Final Reward', 'Best Reward', 'Mean Reward', 'Std Reward',
                          'Final Loss', 'Min Loss', 'Mean Loss', 'Final Epsilon', 'Final Cost (M KRW)',
                          'Final Runoff Reduction (m³)'],
                'Value': [len(episodes), df['Reward'].iloc[-1], df['Reward'].max(), df['Reward'].mean(),
                         df['Reward'].std(), df['Loss'].iloc[-1], df['Loss'].min(), df['Loss'].mean(),
                         df['Epsilon'].iloc[-1], df['Total_Cost_KRW'].iloc[-1] / 1000000,
                         df['Runoff_Reduction_m3'].iloc[-1]]
            })
            summary_stats.to_excel(writer, sheet_name='Summary_Statistics', index=False)
            
            # Format the Excel file
            self._format_excel_file(writer, df, summary_stats)
        
        self.logger.info(f"Training metrics Excel saved: {output_file}")
        return output_file
    
    def create_lid_placement_summary_excel(self, lid_placements: List[Dict], baseline_runoff: float, training_metrics: TrainingMetrics) -> Path:
        """
        Generate LID placement summary Excel file
        
        Args:
            lid_placements: List of LID placement dictionaries
            baseline_runoff: Baseline runoff value (m³)
            training_metrics: Training metrics from agent
            
        Returns:
            Path to generated file
        """
        self.logger.info("Creating LID placement summary Excel file...")
        
        # Prepare placement data
        if lid_placements:
            placement_df = pd.DataFrame(lid_placements)
            
            # Add cost per m² column
            placement_df['Cost_per_m2_KRW'] = placement_df.apply(
                lambda row: LID_COSTS.get(row['lid_type'], 0.0), axis=1
            )
            
            # Rename columns for clarity
            placement_df.rename(columns={
                'lid_type': 'LID_Type',
                'area_m2': 'Area_m2',
                'area_percentage': 'Area_Percentage_of_Subcatchment',
                'cost_krw': 'Total_Cost_KRW'
            }, inplace=True)
        else:
            # Empty DataFrame if no placements
            placement_df = pd.DataFrame(columns=[
                'LID_Type', 'Area_m2', 'Area_Percentage_of_Subcatchment', 
                'Total_Cost_KRW', 'Cost_per_m2_KRW'
            ])
        
        # Add Total row to placement_df
        if not placement_df.empty:
            total_area = placement_df['Area_m2'].sum()
            total_cost = placement_df['Total_Cost_KRW'].sum()
            total_percentage = placement_df['Area_Percentage_of_Subcatchment'].sum()
            avg_cost_per_m2 = total_cost / total_area if total_area > 0 else 0
            
            total_row = pd.DataFrame({
                'LID_Type': ['TOTAL'],
                'Area_m2': [total_area],
                'Area_Percentage_of_Subcatchment': [total_percentage],
                'Total_Cost_KRW': [total_cost],
                'Cost_per_m2_KRW': [avg_cost_per_m2]
            })
            
            placement_df = pd.concat([placement_df, total_row], ignore_index=True)
        
        # Create summary statistics
        if not placement_df.empty:
            # Calculate runoff reduction metrics (exclude TOTAL row for calculations)
            data_df = placement_df[placement_df['LID_Type'] != 'TOTAL']  # Exclude TOTAL row
            
            # Get final runoff reduction from training metrics
            final_runoff_reduction = training_metrics.episode_runoff_reductions[-1] if training_metrics.episode_runoff_reductions else 0.0
            runoff_reduction_percentage = (final_runoff_reduction / baseline_runoff) * 100 if baseline_runoff > 0 else 0.0
            
            # Calculate cost efficiency (exclude TOTAL row)
            total_cost_million = data_df['Total_Cost_KRW'].sum() / 1000000 if not data_df.empty else 0
            runoff_reduction_per_cost = final_runoff_reduction / total_cost_million if total_cost_million > 0 else 0.0
            
            summary_df = pd.DataFrame({
                'Metric': [
                    'Total LID Types Used',
                    'Total LID Area (m²)',
                    'Total LID Cost (KRW)',
                    'Total LID Cost (M KRW)',
                    'Average Cost per m² (KRW)',
                    'Most Expensive LID Type',
                    'Largest LID Area (m²)',
                    'Most Area-Efficient LID Type',
                    'Total Runoff Reduction (m³)',
                    'Total Runoff Reduction (%)',
                    'Runoff Reduction per Cost (m³/M KRW)'
                ],
                'Value': [
                    len(data_df),
                    data_df['Area_m2'].sum(),
                    data_df['Total_Cost_KRW'].sum(),
                    total_cost_million,
                    data_df['Total_Cost_KRW'].sum() / data_df['Area_m2'].sum() if data_df['Area_m2'].sum() > 0 else 0,
                    data_df.loc[data_df['Total_Cost_KRW'].idxmax(), 'LID_Type'] if not data_df.empty else 'N/A',
                    data_df['Area_m2'].max(),
                    data_df.loc[data_df['Area_m2'].idxmax(), 'LID_Type'] if not data_df.empty else 'N/A',
                    final_runoff_reduction,
                    runoff_reduction_percentage,
                    runoff_reduction_per_cost
                ]
            })
        else:
            summary_df = pd.DataFrame({
                'Metric': ['Total LID Types Used'],
                'Value': [0]
            })
        
        # LID costs reference table
        costs_df = pd.DataFrame({
            'LID_Type': list(LID_COSTS.keys()),
            'Cost_per_m2_KRW': list(LID_COSTS.values())
        })
        costs_df = costs_df.sort_values('Cost_per_m2_KRW')
        
        # Save to Excel
        output_file = self.output_dir / 'lid_placement_summary.xlsx'
        
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # LID placements sheet
            placement_df.to_excel(writer, sheet_name='LID_Placements', index=False)
            
            # Summary sheet
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
            
            # LID costs reference sheet
            costs_df.to_excel(writer, sheet_name='LID_Costs_Reference', index=False)
            
            # Format Excel file
            self._format_lid_excel_file(writer, placement_df, summary_df, costs_df)
        
        self.logger.info(f"LID placement summary Excel saved: {output_file}")
        return output_file
    
    def create_baseline_comparison_chart(self, 
                                       training_metrics: TrainingMetrics,
                                       baseline_runoff: float,
                                       lid_placements: List[Dict]) -> Path:
        """
        Generate baseline comparison chart
        
        Args:
            training_metrics: Training metrics
            baseline_runoff: Baseline runoff value
            lid_placements: Final LID placements
            
        Returns:
            Path to generated file
        """
        self.logger.info("Creating baseline comparison chart...")
        
        fig, ((ax1, ax2), (ax3, ax4)) = plt.subplots(2, 2, figsize=(16, 12))
        fig.suptitle('RLID-NET: Baseline vs Final Performance Comparison', fontsize=16, fontweight='bold')
        
        # 1. Runoff comparison (top-left)
        final_runoff_reduction = training_metrics.episode_runoff_reductions[-1] if training_metrics.episode_runoff_reductions else 0
        final_runoff = baseline_runoff - final_runoff_reduction
        
        runoff_data = [baseline_runoff, final_runoff, final_runoff_reduction]
        runoff_labels = ['Baseline\nRunoff', 'Final\nRunoff', 'Runoff\nReduction']
        colors = ['#FF6B6B', '#4ECDC4', '#45B7D1']
        
        bars1 = ax1.bar(runoff_labels, runoff_data, color=colors, alpha=0.8)
        ax1.set_ylabel('Runoff (m³)', fontweight='bold')
        ax1.set_title('Runoff Comparison', fontweight='bold')
        ax1.grid(True, alpha=0.3)
        
        # Add value labels on bars
        for bar, value in zip(bars1, runoff_data):
            height = bar.get_height()
            ax1.text(bar.get_x() + bar.get_width()/2., height,
                    f'{value:.1f}', ha='center', va='bottom', fontweight='bold')
        
        # 2. Cost breakdown (top-right)
        if lid_placements:
            lid_types = [p['lid_type'] for p in lid_placements]
            lid_costs = [p['cost_krw'] / 1000 for p in lid_placements]  # Convert to thousands
            
            wedges, texts, autotexts = ax2.pie(lid_costs, labels=lid_types, autopct='%1.1f%%', startangle=90)
            ax2.set_title('LID Cost Distribution (K KRW)', fontweight='bold')
        else:
            ax2.text(0.5, 0.5, 'No LID\nPlacements', ha='center', va='center', 
                    transform=ax2.transAxes, fontsize=14, fontweight='bold')
            ax2.set_title('LID Cost Distribution', fontweight='bold')
        
        # 3. Training progress (bottom-left)
        episodes = range(1, len(training_metrics.episode_rewards) + 1)
        ax3.plot(episodes, training_metrics.episode_rewards, color='#2E86AB', linewidth=2, label='Reward')
        ax3_twin = ax3.twinx()
        ax3_twin.plot(episodes, training_metrics.episode_runoff_reductions, color='#A23B72', linewidth=2, label='Runoff Reduction')
        
        ax3.set_xlabel('Episode', fontweight='bold')
        ax3.set_ylabel('Reward', color='#2E86AB', fontweight='bold')
        ax3_twin.set_ylabel('Runoff Reduction (m³)', color='#A23B72', fontweight='bold')
        ax3.set_title('Training Progress', fontweight='bold')
        ax3.grid(True, alpha=0.3)
        
        # 4. Performance metrics (bottom-right)
        if lid_placements:
            total_area = sum(p['area_m2'] for p in lid_placements)
            total_cost = sum(p['cost_krw'] for p in lid_placements) / 1000000  # Convert to millions
            reduction_percentage = (final_runoff_reduction / baseline_runoff) * 100 if baseline_runoff > 0 else 0
            
            metrics = ['Total LID\nArea (m²)', 'Total Cost\n(M KRW)', 'Runoff Reduction\n(%)', 'Cost per m³\nReduction (K KRW)']
            values = [
                total_area,
                total_cost,
                reduction_percentage,
                (total_cost * 1000) / final_runoff_reduction if final_runoff_reduction > 0 else 0
            ]
            
            bars4 = ax4.bar(metrics, values, color=['#FF9F43', '#26D0CE', '#45B7D1', '#FD79A8'], alpha=0.8)
            ax4.set_title('Final Performance Metrics', fontweight='bold')
            ax4.grid(True, alpha=0.3)
            
            # Add value labels
            for bar, value in zip(bars4, values):
                height = bar.get_height()
                ax4.text(bar.get_x() + bar.get_width()/2., height,
                        f'{value:.1f}', ha='center', va='bottom', fontweight='bold')
        else:
            ax4.text(0.5, 0.5, 'No Performance\nData Available', ha='center', va='center',
                    transform=ax4.transAxes, fontsize=14, fontweight='bold')
            ax4.set_title('Final Performance Metrics', fontweight='bold')
        
        plt.tight_layout()
        
        # Save figure
        output_file = self.output_dir / 'baseline_comparison.png'
        plt.savefig(output_file, bbox_inches='tight', facecolor='white')
        plt.close()
        
        self.logger.info(f"Baseline comparison chart saved: {output_file}")
        return output_file
    
    def _format_excel_file(self, writer, df: pd.DataFrame, summary_df: pd.DataFrame):
        """Format Excel file with styling"""
        try:
            # Access workbook and worksheets
            workbook = writer.book
            
            # Format training metrics sheet
            if 'Training_Metrics' in workbook.sheetnames:
                ws1 = workbook['Training_Metrics']
                self._apply_excel_formatting(ws1, df)
            
            # Format summary sheet
            if 'Summary_Statistics' in workbook.sheetnames:
                ws2 = workbook['Summary_Statistics']
                self._apply_excel_formatting(ws2, summary_df)
                
        except Exception as e:
            self.logger.warning(f"Could not apply Excel formatting: {str(e)}")
    
    def _format_lid_excel_file(self, writer, placement_df: pd.DataFrame, 
                              summary_df: pd.DataFrame, costs_df: pd.DataFrame):
        """Format LID placement Excel file with styling"""
        try:
            workbook = writer.book
            
            # Format each sheet
            for sheet_name, df in [('LID_Placements', placement_df), 
                                  ('Summary', summary_df), 
                                  ('LID_Costs_Reference', costs_df)]:
                if sheet_name in workbook.sheetnames:
                    ws = workbook[sheet_name]
                    self._apply_excel_formatting(ws, df)
                    
        except Exception as e:
            self.logger.warning(f"Could not apply Excel formatting: {str(e)}")
    
    def _apply_excel_formatting(self, worksheet, df: pd.DataFrame):
        """Apply consistent formatting to Excel worksheet"""
        try:
            # Header formatting
            header_font = Font(bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center')
            
            # Apply header formatting
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
                
        except Exception as e:
            self.logger.warning(f"Could not apply detailed Excel formatting: {str(e)}")


def test_visualizer():
    """Test the visualization system"""
    import tempfile
    from ..rl.agent import TrainingMetrics
    
    logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
    
    # Create test data
    test_metrics = TrainingMetrics(
        episode_rewards=[10.5, 15.2, 12.8, 18.3, 22.1],
        episode_losses=[0.5, 0.3, 0.25, 0.2, 0.15],
        episode_costs=[100000, 150000, 120000, 180000, 200000],
        episode_runoff_reductions=[50.0, 75.0, 60.0, 90.0, 100.0],
        epsilon_values=[1.0, 0.95, 0.90, 0.85, 0.80]
    )
    
    test_placements = [
        {'lid_type': 'Rain Garden', 'area_m2': 100.0, 'area_percentage': 2.0, 'cost_krw': 1500000},
        {'lid_type': 'Green Roof', 'area_m2': 50.0, 'area_percentage': 1.0, 'cost_krw': 8650000}
    ]
    
    # Create visualizer
    with tempfile.TemporaryDirectory() as temp_dir:
        visualizer = RLIDVisualizer(temp_dir)
        
        print("Testing RLID Visualizer:")
        
        # Generate all reports
        files = visualizer.generate_all_reports(
            training_metrics=test_metrics,
            baseline_runoff=500.0,
            final_lid_placements=test_placements
        )
        
        print("Generated files:")
        for file_type, file_path in files.items():
            print(f"   {file_type}: {Path(file_path).name}")
        
        print("Visualizer test completed!")


if __name__ == "__main__":
    test_visualizer() 